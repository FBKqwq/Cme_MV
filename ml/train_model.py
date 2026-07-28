# -*- coding: utf-8 -*-
"""
医院病人“诊断结果”四分类模型训练脚本

目标类别：
    其他、炎症、感染、肿瘤

主要输出：
1. diagnosis_classifier.joblib
2. 模型评估报告.xlsx
3. 测试集预测结果.xlsx
4. 混淆矩阵.png
5. 特征重要性.png
6. 新病人预测模板.xlsx

模型训练成功后，还会自动复制到：
    hospital-diagnosis-system/backend/models/diagnosis_classifier.joblib

首次运行：
    pip install pandas numpy openpyxl scikit-learn joblib matplotlib
"""

from __future__ import annotations

import json
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import joblib
import matplotlib.pyplot as plt
import sklearn
from matplotlib import font_manager


# =========================================================
# 0. 中文字体
# =========================================================
def configure_chinese_font() -> None:
    """尽量选择系统中已经安装的中文字体。"""

    preferred_names = [
        "Microsoft YaHei",
        "SimHei",
        "Noto Sans CJK SC",
        "Noto Sans CJK JP",
        "Arial Unicode MS",
        "AR PL SungtiL GB",
    ]

    available_names = {
        item.name
        for item in font_manager.fontManager.ttflist
    }

    for name in preferred_names:
        if name in available_names:
            plt.rcParams["font.sans-serif"] = [
                name,
                "DejaVu Sans",
            ]
            plt.rcParams["axes.unicode_minus"] = False
            return

    known_paths = [
        Path(
            "/usr/share/fonts/opentype/noto/"
            "NotoSansCJK-Bold.ttc"
        ),
        Path(
            "/usr/share/fonts/truetype/arphic-gbsn00lp/"
            "gbsn00lp.ttf"
        ),
    ]

    for font_path in known_paths:
        if not font_path.exists():
            continue

        try:
            font_manager.fontManager.addfont(
                str(font_path)
            )

            font_name = font_manager.FontProperties(
                fname=str(font_path)
            ).get_name()

            plt.rcParams["font.sans-serif"] = [
                font_name,
                "DejaVu Sans",
            ]
            plt.rcParams["axes.unicode_minus"] = False
            return
        except Exception:
            continue


configure_chinese_font()


import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sklearn.base import clone
from sklearn.compose import ColumnTransformer
from sklearn.dummy import DummyClassifier
from sklearn.ensemble import (
    ExtraTreesClassifier,
    RandomForestClassifier,
)
from sklearn.impute import SimpleImputer
from sklearn.inspection import permutation_importance
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    balanced_accuracy_score,
    classification_report,
    confusion_matrix,
    f1_score,
    log_loss,
    roc_auc_score,
)
from sklearn.model_selection import (
    GridSearchCV,
    StratifiedKFold,
    train_test_split,
)
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import (
    OneHotEncoder,
    StandardScaler,
)


# =========================================================
# 1. 路径和建模配置
# =========================================================
ML_DIR = Path(__file__).resolve().parent

PROJECT_ROOT = ML_DIR.parent

DATA_FILE = (
    ML_DIR
    / "data"
    / "病人档案数据all_清洗后.xlsx"
)

SHEET_NAME = "清洗后数据"

OUTPUT_DIR = (
    ML_DIR
    / "output"
)

BACKEND_MODEL_PATH = (
    PROJECT_ROOT
    / "backend"
    / "models"
    / "diagnosis_classifier.joblib"
)
TARGET = "诊断结果"

LABELS = [
    "其他",
    "炎症",
    "感染",
    "肿瘤",
]

RANDOM_STATE = 42
TEST_SIZE = 0.20

MODEL_PACKAGE_VERSION = "1.0.0"

# 确诊前预测时，不应使用治疗后的信息。
USE_TREATMENT_FEATURES = False

# quick：普通电脑推荐。
# full：搜索参数更多，训练时间更长。
SEARCH_MODE = "quick"


# =========================================================
# 2. 通用工具函数
# =========================================================
def make_one_hot_encoder() -> OneHotEncoder:
    """兼容不同版本的 scikit-learn。"""

    try:
        return OneHotEncoder(
            handle_unknown="ignore",
            sparse_output=False,
        )
    except TypeError:
        return OneHotEncoder(
            handle_unknown="ignore",
            sparse=False,
        )


def style_excel(path: Path) -> None:
    """格式化输出 Excel。"""

    workbook = load_workbook(path)

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for column_cells in worksheet.columns:
            letter = column_cells[0].column_letter
            max_length = 0

            for cell in column_cells[:300]:
                value = (
                    ""
                    if cell.value is None
                    else str(cell.value)
                )

                max_length = max(
                    max_length,
                    len(value),
                )

            worksheet.column_dimensions[
                letter
            ].width = min(
                max(max_length + 2, 10),
                38,
            )

    workbook.save(path)


def save_confusion_matrix(
    matrix: np.ndarray,
    labels: list[str],
    path: Path,
) -> None:
    """保存混淆矩阵图片。"""

    figure, axis = plt.subplots(
        figsize=(7, 6)
    )

    image = axis.imshow(matrix)

    figure.colorbar(
        image,
        ax=axis,
    )

    axis.set_xticks(
        range(len(labels)),
        labels=labels,
    )

    axis.set_yticks(
        range(len(labels)),
        labels=labels,
    )

    axis.set_xlabel("预测类别")
    axis.set_ylabel("真实类别")
    axis.set_title("测试集混淆矩阵")

    threshold = (
        matrix.max() / 2
        if matrix.size
        else 0
    )

    for row_index in range(matrix.shape[0]):
        for column_index in range(
            matrix.shape[1]
        ):
            value = matrix[
                row_index,
                column_index,
            ]

            axis.text(
                column_index,
                row_index,
                str(value),
                ha="center",
                va="center",
                color=(
                    "white"
                    if value > threshold
                    else "black"
                ),
            )

    figure.tight_layout()

    figure.savefig(
        path,
        dpi=180,
        bbox_inches="tight",
    )

    plt.close(figure)


def save_feature_importance(
    importance_df: pd.DataFrame,
    path: Path,
) -> None:
    """保存前 20 个字段的重要性图片。"""

    shown = (
        importance_df
        .head(20)
        .sort_values("重要性均值")
    )

    figure, axis = plt.subplots(
        figsize=(9, 7)
    )

    axis.barh(
        shown["特征"],
        shown["重要性均值"],
    )

    axis.set_xlabel(
        "置换后 Macro-F1 下降量"
    )

    axis.set_title(
        "前20个字段的重要性"
    )

    figure.tight_layout()

    figure.savefig(
        path,
        dpi=180,
        bbox_inches="tight",
    )

    plt.close(figure)


def build_preprocessor(
    features: pd.DataFrame,
) -> tuple[
    ColumnTransformer,
    list[str],
    list[str],
]:
    """创建数值和类别字段预处理流程。"""

    numeric_features = (
        features
        .select_dtypes(include=[np.number])
        .columns
        .tolist()
    )

    categorical_features = (
        features
        .select_dtypes(exclude=[np.number])
        .columns
        .tolist()
    )

    numeric_pipeline = Pipeline(
        steps=[
            (
                "imputer",
                SimpleImputer(
                    strategy="median"
                ),
            ),
            (
                "scaler",
                StandardScaler(),
            ),
        ]
    )

    categorical_pipeline = Pipeline(
        steps=[
            (
                "imputer",
                SimpleImputer(
                    strategy="most_frequent"
                ),
            ),
            (
                "encoder",
                make_one_hot_encoder(),
            ),
        ]
    )

    preprocessor = ColumnTransformer(
        transformers=[
            (
                "numeric",
                numeric_pipeline,
                numeric_features,
            ),
            (
                "categorical",
                categorical_pipeline,
                categorical_features,
            ),
        ],
        remainder="drop",
    )

    return (
        preprocessor,
        numeric_features,
        categorical_features,
    )


def model_grids(
) -> dict[
    str,
    tuple[
        object,
        dict[str, list[object]],
    ],
]:
    """返回需要比较的模型和参数网格。"""

    if SEARCH_MODE == "full":
        return {
            "逻辑回归": (
                LogisticRegression(
                    max_iter=5000,
                    class_weight="balanced",
                    random_state=RANDOM_STATE,
                ),
                {
                    "model__C": [
                        0.03,
                        0.1,
                        0.3,
                        1.0,
                        3.0,
                        10.0,
                    ],
                },
            ),
            "随机森林": (
                RandomForestClassifier(
                    n_estimators=600,
                    class_weight=(
                        "balanced_subsample"
                    ),
                    random_state=RANDOM_STATE,
                    n_jobs=1,
                ),
                {
                    "model__max_depth": [
                        None,
                        6,
                        10,
                        15,
                    ],
                    "model__min_samples_leaf": [
                        1,
                        2,
                        4,
                        8,
                    ],
                    "model__max_features": [
                        "sqrt",
                        0.5,
                    ],
                },
            ),
            "极端随机树": (
                ExtraTreesClassifier(
                    n_estimators=600,
                    class_weight="balanced",
                    random_state=RANDOM_STATE,
                    n_jobs=1,
                ),
                {
                    "model__max_depth": [
                        None,
                        6,
                        10,
                        15,
                    ],
                    "model__min_samples_leaf": [
                        1,
                        2,
                        4,
                        8,
                    ],
                    "model__max_features": [
                        "sqrt",
                        0.5,
                    ],
                },
            ),
        }

    return {
        "逻辑回归": (
            LogisticRegression(
                max_iter=5000,
                class_weight="balanced",
                random_state=RANDOM_STATE,
            ),
            {
                "model__C": [
                    0.1,
                    1.0,
                    10.0,
                ],
            },
        ),
        "随机森林": (
            RandomForestClassifier(
                n_estimators=400,
                class_weight=(
                    "balanced_subsample"
                ),
                random_state=RANDOM_STATE,
                n_jobs=1,
            ),
            {
                "model__max_depth": [
                    None,
                    10,
                ],
                "model__min_samples_leaf": [
                    1,
                    4,
                ],
                "model__max_features": [
                    "sqrt",
                    0.5,
                ],
            },
        ),
        "极端随机树": (
            ExtraTreesClassifier(
                n_estimators=400,
                class_weight="balanced",
                random_state=RANDOM_STATE,
                n_jobs=1,
            ),
            {
                "model__max_depth": [
                    None,
                    10,
                ],
                "model__min_samples_leaf": [
                    1,
                    4,
                ],
                "model__max_features": [
                    "sqrt",
                    0.5,
                ],
            },
        ),
    }


def create_prediction_template(
    features: pd.DataFrame,
    categorical_features: list[str],
    output_path: Path,
) -> None:
    """创建新病人输入模板。"""

    template = pd.DataFrame(
        columns=features.columns
    )

    instructions: list[
        dict[str, str]
    ] = []

    for column in features.columns:
        if column in categorical_features:
            available_values = sorted(
                features[column]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )

            instructions.append(
                {
                    "字段": column,
                    "类型": "类别",
                    "可参考的取值": " / ".join(
                        available_values
                    ),
                    "说明": (
                        "请使用与训练数据一致的文字"
                    ),
                }
            )
        else:
            instructions.append(
                {
                    "字段": column,
                    "类型": "数值",
                    "可参考的取值": "",
                    "说明": (
                        "填写数字；未知可留空，"
                        "模型会按训练规则填补"
                    ),
                }
            )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        template.to_excel(
            writer,
            sheet_name="新病人数据",
            index=False,
        )

        pd.DataFrame(
            instructions
        ).to_excel(
            writer,
            sheet_name="填写说明",
            index=False,
        )

    style_excel(output_path)


def create_feature_schema(
    features: pd.DataFrame,
    numeric_features: list[str],
    categorical_features: list[str],
) -> list[dict[str, Any]]:
    """
    创建后端和前端后续可以使用的字段说明。
    """

    schema: list[dict[str, Any]] = []

    for column in features.columns:
        if column in numeric_features:
            numeric_series = pd.to_numeric(
                features[column],
                errors="coerce",
            )

            non_missing = (
                numeric_series.dropna()
            )

            schema.append(
                {
                    "name": column,
                    "type": "numeric",
                    "required": False,
                    "min": (
                        float(non_missing.min())
                        if not non_missing.empty
                        else None
                    ),
                    "max": (
                        float(non_missing.max())
                        if not non_missing.empty
                        else None
                    ),
                    "median": (
                        float(non_missing.median())
                        if not non_missing.empty
                        else None
                    ),
                }
            )
        elif column in categorical_features:
            values = sorted(
                features[column]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )

            schema.append(
                {
                    "name": column,
                    "type": "categorical",
                    "required": False,
                    "allowed_values": values,
                }
            )

    return schema


def verify_model_package(
    model_path: Path,
    sample_features: pd.DataFrame,
) -> dict[str, Any]:
    """
    重新加载模型文件并执行一次预测，
    防止保存出空文件或损坏文件。
    """

    if not model_path.exists():
        raise FileNotFoundError(
            f"模型文件不存在：{model_path}"
        )

    file_size = model_path.stat().st_size

    if file_size <= 0:
        raise RuntimeError(
            f"模型文件为 0 字节：{model_path}"
        )

    package = joblib.load(model_path)

    if not isinstance(package, dict):
        raise TypeError(
            "模型文件内容不是模型包字典。"
        )

    required_keys = {
        "model",
        "feature_columns",
        "numeric_features",
        "categorical_features",
        "class_labels",
        "best_model_name",
        "test_metrics",
        "created_at",
    }

    missing_keys = (
        required_keys - set(package.keys())
    )

    if missing_keys:
        raise KeyError(
            "模型包缺少必要内容："
            f"{sorted(missing_keys)}"
        )

    model = package["model"]

    if not hasattr(model, "predict"):
        raise TypeError(
            "模型不支持 predict。"
        )

    if not hasattr(
        model,
        "predict_proba",
    ):
        raise TypeError(
            "模型不支持 predict_proba。"
        )

    feature_columns = list(
        package["feature_columns"]
    )

    if not feature_columns:
        raise ValueError(
            "模型包中的特征列表为空。"
        )

    sample = (
        sample_features
        .reindex(columns=feature_columns)
        .head(1)
        .copy()
    )

    if sample.empty:
        raise ValueError(
            "没有可用于模型验证的样本。"
        )

    prediction = model.predict(sample)

    probabilities = (
        model.predict_proba(sample)
    )

    if len(prediction) != 1:
        raise RuntimeError(
            "模型验证预测结果数量异常。"
        )

    if probabilities.shape[0] != 1:
        raise RuntimeError(
            "模型验证概率结果数量异常。"
        )

    if not np.all(
        np.isfinite(probabilities)
    ):
        raise RuntimeError(
            "模型输出概率中包含无效数值。"
        )

    probability_sum = float(
        probabilities[0].sum()
    )

    if not np.isclose(
        probability_sum,
        1.0,
        atol=1e-5,
    ):
        raise RuntimeError(
            "模型类别概率之和不等于 1。"
        )

    predicted_label = str(
        prediction[0]
    )

    if predicted_label not in set(
        package["class_labels"]
    ):
        raise RuntimeError(
            "模型预测结果不在规定类别中："
            f"{predicted_label}"
        )

    return package


def save_model_package(
    model_package: dict[str, Any],
    output_model_path: Path,
    backend_model_path: Path,
    sample_features: pd.DataFrame,
) -> dict[str, Any]:
    """
    保存模型到临时文件，验证后再替换正式文件，
    最后复制到 FastAPI 后端。
    """

    output_model_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    backend_model_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    temporary_model_path = (
        output_model_path
        .with_suffix(".tmp.joblib")
    )

    if temporary_model_path.exists():
        temporary_model_path.unlink()

    print()
    print("正在保存模型临时文件……")

    joblib.dump(
        model_package,
        temporary_model_path,
        compress=3,
    )

    temporary_size = (
        temporary_model_path.stat().st_size
    )

    if temporary_size <= 0:
        raise RuntimeError(
            "模型保存失败：临时模型为 0 字节。"
        )

    print(
        "正在重新加载并验证临时模型……"
    )

    verified_package = verify_model_package(
        temporary_model_path,
        sample_features,
    )

    # os.replace 语义，目标存在时也会替换。
    temporary_model_path.replace(
        output_model_path
    )

    print(
        "正在复制模型到 FastAPI 后端……"
    )

    shutil.copy2(
        output_model_path,
        backend_model_path,
    )

    backend_package = verify_model_package(
        backend_model_path,
        sample_features,
    )

    output_size = (
        output_model_path.stat().st_size
    )

    backend_size = (
        backend_model_path.stat().st_size
    )

    if output_size != backend_size:
        raise RuntimeError(
            "训练输出模型和后端模型大小不一致。"
        )

    print()
    print("模型文件保存并验证成功：")
    print(
        f"训练输出模型：{output_model_path}"
    )
    print(
        f"训练输出大小：{output_size:,} 字节"
    )
    print(
        f"后端模型：{backend_model_path}"
    )
    print(
        f"后端模型大小：{backend_size:,} 字节"
    )
    print(
        "模型对象类型："
        f"{type(backend_package['model'])}"
    )
    print(
        "模型输入字段数："
        f"{len(backend_package['feature_columns'])}"
    )
    print(
        "模型类别："
        f"{backend_package['class_labels']}"
    )

    return verified_package


# =========================================================
# 3. 主训练流程
# =========================================================
def main() -> None:
    # -----------------------------------------------------
    # 3.1 检查并读取数据
    # -----------------------------------------------------
    if not DATA_FILE.exists():
        raise FileNotFoundError(
            f"找不到数据文件：{DATA_FILE}\n"
            "请检查 DATA_FILE 是否正确。"
        )

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    print("=" * 60)
    print("诊断结果四分类模型训练")
    print("=" * 60)
    print(f"数据文件：{DATA_FILE}")
    print(f"工作表：{SHEET_NAME}")
    print(f"输出目录：{OUTPUT_DIR}")
    print()

    excel_file = pd.ExcelFile(DATA_FILE)

    if SHEET_NAME not in excel_file.sheet_names:
        raise KeyError(
            f"Excel 中不存在工作表：{SHEET_NAME}\n"
            f"实际工作表：{excel_file.sheet_names}"
        )

    data = pd.read_excel(
        DATA_FILE,
        sheet_name=SHEET_NAME,
    )

    data.columns = (
        data.columns
        .astype(str)
        .str.strip()
    )

    if TARGET not in data.columns:
        raise KeyError(
            f"数据中不存在目标列：{TARGET}"
        )

    missing_markers = [
        "",
        " ",
        "未查",
        "未知",
        "NA",
        "N/A",
        "null",
        "NULL",
        "-",
        "--",
    ]

    data = data.replace(
        missing_markers,
        np.nan,
    )

    missing_target_count = int(
        data[TARGET].isna().sum()
    )

    if missing_target_count > 0:
        print(
            f"发现 {missing_target_count} 条"
            "诊断结果为空的记录，训练时自动排除。"
        )

        data = data.loc[
            data[TARGET].notna()
        ].copy()

    data[TARGET] = (
        data[TARGET]
        .astype(str)
        .str.strip()
    )

    unexpected_labels = sorted(
        set(data[TARGET]) - set(LABELS)
    )

    missing_labels = sorted(
        set(LABELS) - set(data[TARGET])
    )

    if unexpected_labels:
        raise ValueError(
            "诊断结果中出现未知类别："
            f"{unexpected_labels}"
        )

    if missing_labels:
        raise ValueError(
            "诊断结果缺少类别："
            f"{missing_labels}"
        )

    # -----------------------------------------------------
    # 3.2 排除泄漏字段和身份字段
    # -----------------------------------------------------
    excluded_columns = [
        TARGET,
        "ID",
        "姓名",
        "感染分类",
        "group",
        "CODE",
        "发烧时长_异常标记",
    ]

    excluded_columns += [
        column
        for column in data.columns
        if column.endswith("_原缺失")
    ]

    treatment_columns = [
        "抗生素",
        "非甾体抗炎药",
        "糖皮质激素",
    ]

    if not USE_TREATMENT_FEATURES:
        excluded_columns += treatment_columns

    excluded_columns = list(
        dict.fromkeys(
            column
            for column in excluded_columns
            if column in data.columns
        )
    )

    features = data.drop(
        columns=excluded_columns
    ).copy()

    target = data[TARGET].copy()

    if features.empty:
        raise ValueError(
            "排除字段后没有可用于建模的特征。"
        )

    # 将正负无穷替换为空值，由模型管道统一填补。
    features = features.replace(
        [np.inf, -np.inf],
        np.nan,
    )

    # 删除所有值完全相同的字段。
    constant_columns = [
        column
        for column in features.columns
        if (
            features[column]
            .nunique(dropna=False)
            <= 1
        )
    ]

    if constant_columns:
        print(
            "删除没有变化的字段："
            f"{constant_columns}"
        )

        features = features.drop(
            columns=constant_columns
        )

        excluded_columns += (
            constant_columns
        )

    if features.empty:
        raise ValueError(
            "删除常量字段后没有可用特征。"
        )

    (
        preprocessor,
        numeric_features,
        categorical_features,
    ) = build_preprocessor(features)

    # -----------------------------------------------------
    # 3.3 划分训练集和测试集
    # -----------------------------------------------------
    (
        X_train,
        X_test,
        y_train,
        y_test,
    ) = train_test_split(
        features,
        target,
        test_size=TEST_SIZE,
        stratify=target,
        random_state=RANDOM_STATE,
    )

    minimum_train_class_count = int(
        y_train.value_counts().min()
    )

    cv_splits = min(
        5,
        minimum_train_class_count,
    )

    if cv_splits < 2:
        raise ValueError(
            "训练集中某个类别的样本数不足，"
            "无法进行分层交叉验证。"
        )

    cross_validation = StratifiedKFold(
        n_splits=cv_splits,
        shuffle=True,
        random_state=RANDOM_STATE,
    )

    print(f"有效样本数：{len(data)}")
    print(
        f"建模字段数：{features.shape[1]}"
    )
    print(
        f"数值字段数：{len(numeric_features)}"
    )
    print(
        "类别字段数："
        f"{len(categorical_features)}"
    )
    print(
        f"交叉验证折数：{cv_splits}"
    )
    print()
    print("目标类别分布：")
    print(
        target
        .value_counts()
        .reindex(LABELS)
    )

    # -----------------------------------------------------
    # 4. 模型比较和参数搜索
    # -----------------------------------------------------
    search_results: dict[
        str,
        GridSearchCV,
    ] = {}

    comparison_rows: list[
        dict[str, object]
    ] = []

    for (
        model_name,
        (
            estimator,
            parameter_grid,
        ),
    ) in model_grids().items():
        pipeline = Pipeline(
            steps=[
                (
                    "preprocessor",
                    clone(preprocessor),
                ),
                (
                    "model",
                    estimator,
                ),
            ]
        )

        search = GridSearchCV(
            estimator=pipeline,
            param_grid=parameter_grid,
            scoring="f1_macro",
            cv=cross_validation,
            n_jobs=-1,
            refit=True,
            return_train_score=False,
            error_score="raise",
        )

        print()
        print(f"正在训练：{model_name}")

        search.fit(
            X_train,
            y_train,
        )

        search_results[
            model_name
        ] = search

        comparison_rows.append(
            {
                "模型": model_name,
                "交叉验证Macro-F1": (
                    search.best_score_
                ),
                "最佳参数": json.dumps(
                    search.best_params_,
                    ensure_ascii=False,
                ),
            }
        )

        print(
            "最佳交叉验证 Macro-F1："
            f"{search.best_score_:.4f}"
        )
        print(
            f"最佳参数：{search.best_params_}"
        )

    comparison_df = (
        pd.DataFrame(comparison_rows)
        .sort_values(
            "交叉验证Macro-F1",
            ascending=False,
        )
        .reset_index(drop=True)
    )

    best_model_name = str(
        comparison_df.iloc[0]["模型"]
    )

    best_estimator = (
        search_results[
            best_model_name
        ].best_estimator_
    )

    # -----------------------------------------------------
    # 5. 测试集评估
    # -----------------------------------------------------
    predicted_labels = (
        best_estimator.predict(X_test)
    )

    predicted_probabilities = (
        best_estimator.predict_proba(
            X_test
        )
    )

    model_classes = (
        best_estimator
        .classes_
        .tolist()
    )

    class_to_index = {
        class_name: index
        for index, class_name
        in enumerate(model_classes)
    }

    missing_probability_classes = (
        set(LABELS)
        - set(model_classes)
    )

    if missing_probability_classes:
        raise RuntimeError(
            "模型缺少类别概率："
            f"{sorted(missing_probability_classes)}"
        )

    test_accuracy = accuracy_score(
        y_test,
        predicted_labels,
    )

    test_balanced_accuracy = (
        balanced_accuracy_score(
            y_test,
            predicted_labels,
        )
    )

    test_macro_f1 = f1_score(
        y_test,
        predicted_labels,
        average="macro",
    )

    test_weighted_f1 = f1_score(
        y_test,
        predicted_labels,
        average="weighted",
    )

    test_log_loss = log_loss(
        y_test,
        predicted_probabilities,
        labels=model_classes,
    )

    probability_for_auc = np.column_stack(
        [
            predicted_probabilities[
                :,
                class_to_index[class_name],
            ]
            for class_name in LABELS
        ]
    )

    try:
        test_auc_macro = roc_auc_score(
            y_test,
            probability_for_auc,
            labels=LABELS,
            average="macro",
            multi_class="ovr",
        )
    except ValueError as error:
        print(
            "警告：测试集 ROC-AUC 无法计算："
            f"{error}"
        )
        test_auc_macro = float("nan")

    test_metrics = pd.DataFrame(
        [
            {
                "指标": "Accuracy",
                "数值": test_accuracy,
            },
            {
                "指标": "Balanced Accuracy",
                "数值": (
                    test_balanced_accuracy
                ),
            },
            {
                "指标": "Macro-F1",
                "数值": test_macro_f1,
            },
            {
                "指标": "Weighted-F1",
                "数值": test_weighted_f1,
            },
            {
                "指标": "Macro ROC-AUC (OvR)",
                "数值": test_auc_macro,
            },
            {
                "指标": "Log Loss",
                "数值": test_log_loss,
            },
        ]
    )

    # -----------------------------------------------------
    # 5.1 多数类基线
    # -----------------------------------------------------
    dummy_model = DummyClassifier(
        strategy="most_frequent"
    )

    dummy_model.fit(
        X_train,
        y_train,
    )

    dummy_prediction = (
        dummy_model.predict(X_test)
    )

    majority_label = str(
        y_train.value_counts().idxmax()
    )

    baseline_comparison = pd.DataFrame(
        [
            {
                "模型": (
                    "多数类基线"
                    f"（始终预测{majority_label}）"
                ),
                "Accuracy": accuracy_score(
                    y_test,
                    dummy_prediction,
                ),
                "Balanced Accuracy": (
                    balanced_accuracy_score(
                        y_test,
                        dummy_prediction,
                    )
                ),
                "Macro-F1": f1_score(
                    y_test,
                    dummy_prediction,
                    average="macro",
                ),
            },
            {
                "模型": best_model_name,
                "Accuracy": test_accuracy,
                "Balanced Accuracy": (
                    test_balanced_accuracy
                ),
                "Macro-F1": test_macro_f1,
            },
        ]
    )

    # -----------------------------------------------------
    # 5.2 分类报告和混淆矩阵
    # -----------------------------------------------------
    report_dict = classification_report(
        y_test,
        predicted_labels,
        labels=LABELS,
        output_dict=True,
        zero_division=0,
    )

    class_report_df = (
        pd.DataFrame(report_dict)
        .T
        .reset_index()
        .rename(
            columns={
                "index": "类别或汇总"
            }
        )
    )

    confusion = confusion_matrix(
        y_test,
        predicted_labels,
        labels=LABELS,
    )

    confusion_df = pd.DataFrame(
        confusion,
        index=[
            f"真实_{label}"
            for label in LABELS
        ],
        columns=[
            f"预测_{label}"
            for label in LABELS
        ],
    ).reset_index(
        names="真实类别"
    )

    # -----------------------------------------------------
    # 5.3 测试集逐条预测结果
    # -----------------------------------------------------
    prediction_df = X_test.copy()

    prediction_df.insert(
        0,
        "原数据行号",
        X_test.index + 2,
    )

    prediction_df.insert(
        1,
        "真实诊断结果",
        y_test.to_numpy(),
    )

    prediction_df.insert(
        2,
        "预测诊断结果",
        predicted_labels,
    )

    prediction_df.insert(
        3,
        "预测是否正确",
        np.where(
            predicted_labels
            == y_test.to_numpy(),
            "是",
            "否",
        ),
    )

    for class_name in LABELS:
        prediction_df[
            f"概率_{class_name}"
        ] = predicted_probabilities[
            :,
            class_to_index[class_name],
        ]

    probability_columns = [
        f"概率_{class_name}"
        for class_name in LABELS
    ]

    prediction_df[
        "最高预测概率"
    ] = prediction_df[
        probability_columns
    ].max(axis=1)

    # -----------------------------------------------------
    # 5.4 字段置换重要性
    # -----------------------------------------------------
    print()
    print("正在计算字段置换重要性……")

    permutation = permutation_importance(
        best_estimator,
        X_test,
        y_test,
        scoring="f1_macro",
        n_repeats=20,
        random_state=RANDOM_STATE,
        n_jobs=-1,
    )

    importance_df = pd.DataFrame(
        {
            "特征": features.columns,
            "重要性均值": (
                permutation.importances_mean
            ),
            "重要性标准差": (
                permutation.importances_std
            ),
        }
    ).sort_values(
        "重要性均值",
        ascending=False,
    )

    class_distribution = (
        target
        .value_counts()
        .reindex(LABELS)
        .rename_axis("诊断结果")
        .reset_index(name="样本数")
    )

    class_distribution[
        "占比"
    ] = (
        class_distribution["样本数"]
        / len(target)
    )

    feature_info = pd.DataFrame(
        {
            "特征": features.columns,
            "字段类型": [
                (
                    "数值"
                    if column
                    in numeric_features
                    else "类别"
                )
                for column
                in features.columns
            ],
            "缺失数": [
                int(
                    features[
                        column
                    ].isna().sum()
                )
                for column
                in features.columns
            ],
            "唯一值数": [
                int(
                    features[
                        column
                    ].nunique(
                        dropna=True
                    )
                )
                for column
                in features.columns
            ],
        }
    )

    excluded_df = pd.DataFrame(
        {
            "已排除字段": (
                excluded_columns
            )
        }
    )

    comparison_df[
        "测试集入选模型"
    ] = np.where(
        comparison_df["模型"]
        == best_model_name,
        "是",
        "否",
    )

    # -----------------------------------------------------
    # 5.5 保存评估报告
    # -----------------------------------------------------
    evaluation_path = (
        OUTPUT_DIR
        / "模型评估报告.xlsx"
    )

    with pd.ExcelWriter(
        evaluation_path,
        engine="openpyxl",
    ) as writer:
        comparison_df.to_excel(
            writer,
            sheet_name="模型比较",
            index=False,
        )

        test_metrics.to_excel(
            writer,
            sheet_name="测试集总体指标",
            index=False,
        )

        baseline_comparison.to_excel(
            writer,
            sheet_name="与多数类基线比较",
            index=False,
        )

        class_report_df.to_excel(
            writer,
            sheet_name="各类别指标",
            index=False,
        )

        confusion_df.to_excel(
            writer,
            sheet_name="混淆矩阵",
            index=False,
        )

        class_distribution.to_excel(
            writer,
            sheet_name="类别分布",
            index=False,
        )

        importance_df.to_excel(
            writer,
            sheet_name="特征重要性",
            index=False,
        )

        feature_info.to_excel(
            writer,
            sheet_name="建模特征说明",
            index=False,
        )

        excluded_df.to_excel(
            writer,
            sheet_name="排除字段",
            index=False,
        )

    style_excel(evaluation_path)

    prediction_path = (
        OUTPUT_DIR
        / "测试集预测结果.xlsx"
    )

    prediction_df.to_excel(
        prediction_path,
        index=False,
    )

    style_excel(prediction_path)

    save_confusion_matrix(
        confusion,
        LABELS,
        OUTPUT_DIR / "混淆矩阵.png",
    )

    save_feature_importance(
        importance_df,
        OUTPUT_DIR / "特征重要性.png",
    )

    # -----------------------------------------------------
    # 6. 使用全部数据重新训练最终模型
    # -----------------------------------------------------
    print()
    print(
        "正在使用全部数据重新训练最终模型……"
    )

    final_model = clone(
        best_estimator
    )

    final_model.fit(
        features,
        target,
    )

    feature_schema = create_feature_schema(
        features,
        numeric_features,
        categorical_features,
    )

    model_package: dict[str, Any] = {
        "package_version": (
            MODEL_PACKAGE_VERSION
        ),
        "model": final_model,
        "feature_columns": (
            features.columns.tolist()
        ),
        "feature_schema": feature_schema,
        "numeric_features": (
            numeric_features
        ),
        "categorical_features": (
            categorical_features
        ),
        "class_labels": LABELS,
        "model_classes": (
            final_model
            .classes_
            .tolist()
        ),
        "excluded_columns": (
            excluded_columns
        ),
        "best_model_name": (
            best_model_name
        ),
        "best_params": (
            search_results[
                best_model_name
            ].best_params_
        ),
        "test_metrics": {
            "accuracy": float(
                test_accuracy
            ),
            "balanced_accuracy": float(
                test_balanced_accuracy
            ),
            "macro_f1": float(
                test_macro_f1
            ),
            "weighted_f1": float(
                test_weighted_f1
            ),
            "macro_roc_auc_ovr": (
                float(test_auc_macro)
            ),
            "log_loss": float(
                test_log_loss
            ),
        },
        "python_version": sys.version,
        "sklearn_version": (
            sklearn.__version__
        ),
        "joblib_version": (
            joblib.__version__
        ),
        "numpy_version": (
            np.__version__
        ),
        "pandas_version": (
            pd.__version__
        ),
        "created_at": datetime.now().isoformat(
            timespec="seconds"
        ),
        "data_file": str(DATA_FILE),
        "sheet_name": SHEET_NAME,
        "sample_count": int(
            len(features)
        ),
        "prediction_time_assumption": (
            "确诊前预测；治疗字段已排除"
            if not USE_TREATMENT_FEATURES
            else "允许使用治疗字段"
        ),
    }

    model_path = (
        OUTPUT_DIR
        / "diagnosis_classifier.joblib"
    )

    save_model_package(
        model_package=model_package,
        output_model_path=model_path,
        backend_model_path=(
            BACKEND_MODEL_PATH
        ),
        sample_features=features,
    )

    # -----------------------------------------------------
    # 7. 保存新病人输入模板
    # -----------------------------------------------------
    template_path = (
        OUTPUT_DIR
        / "新病人预测模板.xlsx"
    )

    create_prediction_template(
        features,
        categorical_features,
        template_path,
    )

    # -----------------------------------------------------
    # 8. 最终输出
    # -----------------------------------------------------
    print()
    print("=" * 60)
    print("建模完成")
    print("=" * 60)
    print(f"入选模型：{best_model_name}")
    print(
        "测试集 Accuracy："
        f"{test_accuracy:.4f}"
    )
    print(
        "测试集 Balanced Accuracy："
        f"{test_balanced_accuracy:.4f}"
    )
    print(
        "测试集 Macro-F1："
        f"{test_macro_f1:.4f}"
    )

    if np.isfinite(test_auc_macro):
        print(
            "测试集 Macro ROC-AUC："
            f"{test_auc_macro:.4f}"
        )
    else:
        print(
            "测试集 Macro ROC-AUC：无法计算"
        )

    print(f"输出目录：{OUTPUT_DIR}")
    print(
        f"后端模型路径：{BACKEND_MODEL_PATH}"
    )
    print("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print()
        print(
            f"程序运行失败：{error}",
            file=sys.stderr,
        )
        raise